#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Frontend Desktop Ultra-Moderno para o Tradutor de Dados Universal
Interface redesenhada com CustomTkinter seguindo as melhores práticas de UI/UX
"""

import customtkinter as ctk
import pandas as pd
import sqlite3
import os
import threading
import queue
import time
from datetime import datetime
from deep_translator import GoogleTranslator
import json
from pathlib import Path
from tkinter import filedialog, messagebox

class TradutorCustomTkinterUX:
    def __init__(self):
        # Configurar aparência do CustomTkinter
        ctk.set_appearance_mode("dark")
        # Não usar tema padrão para ter controle total das cores
        
        # Criar janela principal
        self.root = ctk.CTk()
        self.root.title("📊 Tradutor de Dados Universal")
        
        # Forçar cor de fundo preta
        self.root.configure(fg_color="#000000")
        
        # Permitir redimensionamento
        self.root.resizable(True, True)

        # Ajustar tamanho inicial dinamicamente à tela do usuário
        self.ajustar_tamanho_para_tela()
        
        # Variáveis de estado
        self.df = None  # Mantido para compatibilidade
        self.df_preview = None  # Preview dos dados (primeiras 50 linhas)
        self.df_full_path = None  # Caminho completo do arquivo
        self.df_tipo = None  # Tipo do arquivo (CSV, Excel, SQLite)
        self.df_tabela = None  # Nome da tabela (para SQLite)
        self.colunas_originais = []
        self.colunas_traduzidas = []
        self.traducao_ativa = False
        self.thread_traducao = None
        self.progress_queue = queue.Queue()
        
        # Configurações
        self.config = {
            'idioma_origem': 'en',
            'idioma_destino': 'pt',
            'tamanho_lote': 15,  # Lotes menores para economizar RAM
            'delay_traducao': 0.3  # Delay menor para melhor responsividade
        }
        
        # Cores e estilos para efeitos visuais - Tema Minimalista Dark
        self.cores = {
            'primary': '#ffffff',  # Branco puro para elementos principais
            'primary_hover': '#e5e7eb',  # Cinza muito claro no hover
            'secondary': '#0f0f0f',  # Preto quase puro para frames secundários
            'secondary_hover': '#1a1a1a',  # Preto mais claro no hover
            'accent': '#059669',  # Verde mais escuro e sutil
            'accent_hover': '#047857',  # Verde ainda mais escuro no hover
            'warning': '#dc2626',  # Vermelho mais escuro
            'warning_hover': '#b91c1c',  # Vermelho ainda mais escuro no hover
            'danger': '#dc2626',
            'danger_hover': '#b91c1c',
            'transparent': 'transparent',
            'semi_transparent': '#000000',  # Preto puro
            'glass': '#0f0f0f',  # Preto quase puro para frames
            'text_primary': '#ffffff',  # Texto principal branco
            'text_secondary': '#d1d5db',  # Texto secundário cinza claro
            'text_muted': '#9ca3af',  # Texto mudo cinza médio
            'border': '#1a1a1a'  # Bordas muito escuras
        }
        
        # Criar interface
        self.criar_interface()
        
        # Configurar estado inicial da tabela SQLite
        self.atualizar_visibilidade_tabela_sqlite()
        
        # Iniciar monitoramento de progresso
        self.monitorar_progresso()
    
    def centralizar_janela(self):
        """Centraliza a janela na tela"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def ajustar_tamanho_para_tela(self):
        """Define um tamanho inicial que caiba na tela atual e centraliza a janela."""
        self.root.update_idletasks()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()

        # Tamanho desejado padrão e mínimos para boa usabilidade
        desired_w, desired_h = 1200, 800
        min_w, min_h = 960, 600

        # Margem para evitar que encoste nas bordas
        margin = 60

        win_w = min(desired_w, max(min_w, screen_w - margin))
        win_h = min(desired_h, max(min_h, screen_h - margin))

        # Centralizar
        pos_x = max(0, (screen_w - win_w) // 2)
        pos_y = max(0, (screen_h - win_h) // 2)

        self.root.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")
        # minsize não deve ultrapassar o tamanho da tela
        self.root.minsize(min(win_w, desired_w), min(win_h, desired_h))
    
    def criar_interface(self):
        """Cria a interface principal com layout otimizado e responsivo"""
        # Configurar grid principal otimizado para melhor distribuição de espaço
        self.root.grid_columnconfigure(0, weight=1, minsize=280)   # Sidebar esquerda (mínimo 280px)
        self.root.grid_columnconfigure(1, weight=2, minsize=500)   # Área central (mínimo 500px)
        self.root.grid_columnconfigure(2, weight=1, minsize=250)   # Sidebar direita (mínimo 250px)
        self.root.grid_rowconfigure(0, weight=1, minsize=400)      # Conteúdo principal (expansível)
        self.root.grid_rowconfigure(1, weight=0, minsize=25)       # Barra de status (altura fixa menor)
        
        # Sidebar esquerda
        self.criar_sidebar_esquerda()
        
        # Área central
        self.criar_area_central()
        
        # Sidebar direita
        self.criar_sidebar_direita()
        
        # Barra de status
        self.criar_barra_status()
    
    # Header removido para dar mais espaço às funcionalidades principais
    
    def criar_sidebar_esquerda(self):
        """Cria a sidebar esquerda com todas as opções de carregamento de dataset - versão ultra-compacta"""
        sidebar_frame = ctk.CTkFrame(self.root, corner_radius=10, fg_color=self.cores['glass'])
        sidebar_frame.grid(row=0, column=0, sticky="nsew", padx=(12, 6), pady=(8, 12))  # Aproveitar espaço do header removido
        
        # Título da sidebar - mais compacto
        sidebar_title = ctk.CTkLabel(
            sidebar_frame,
            text="⚙️ Configurações",
            font=ctk.CTkFont(size=13, weight="bold"),  # Reduzir de 16 para 13
            text_color=self.cores['text_primary']
        )
        sidebar_title.pack(pady=8)  # Reduzir de 12 para 8
        
        # Container scrollável para configurações
        config_container = ctk.CTkScrollableFrame(
            sidebar_frame, 
            fg_color="transparent"
        )
        config_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))  # Reduzir padding
        
        # Seleção de tipo de arquivo
        self.criar_selecao_tipo_arquivo(config_container)
        
        # Seleção de idiomas
        self.criar_selecao_idiomas(config_container)
        
        # Configurações avançadas
        self.criar_configuracoes_avancadas(config_container)
        
        # Botões de ação
        self.criar_botoes_acao(config_container)
    
    def criar_selecao_tipo_arquivo(self, parent):
        """Cria a seleção de tipo de arquivo com layout organizado - versão compacta"""
        # Frame para tipo de arquivo - mais compacto
        tipo_frame = ctk.CTkFrame(parent, corner_radius=8, fg_color=self.cores['semi_transparent'])
        tipo_frame.pack(fill="x", pady=(0, 8))  # Reduzir espaçamento
        
        # Título do frame - mais compacto
        ctk.CTkLabel(
            tipo_frame,
            text="📋 Tipo de Arquivo",
            font=ctk.CTkFont(size=11, weight="bold"),  # Reduzir de 14 para 11
            text_color=self.cores['text_primary']
        ).pack(pady=6)  # Reduzir de 10 para 6
        
        # Container para tipos - mais compacto
        tipos_container = ctk.CTkFrame(tipo_frame, fg_color="transparent")
        tipos_container.pack(fill="x", padx=10, pady=(0, 8))  # Reduzir padding
        
        self.tipo_arquivo = ctk.StringVar(value="CSV")
        tipos = [("CSV", "CSV"), ("Excel", "Excel"), ("SQLite", "SQLite")]
        
        for text, value in tipos:
            radio = ctk.CTkRadioButton(
                tipos_container,
                text=text,
                variable=self.tipo_arquivo,
                value=value,
                font=ctk.CTkFont(size=10),  # Reduzir de 12 para 10
                fg_color=self.cores['accent'],
                hover_color=self.cores['accent_hover'],
                radiobutton_width=14,  # Reduzir de 16 para 14
                radiobutton_height=14,  # Reduzir de 16 para 14
                text_color=self.cores['text_primary'],
                command=self.atualizar_visibilidade_tabela_sqlite
            )
            radio.pack(anchor="w", pady=(0, 2))  # Reduzir espaçamento
        
        # Botão selecionar arquivo (responsivo) - mais compacto
        btn_selecionar = ctk.CTkButton(
            tipos_container,
            text="Selecionar Arquivo",
            font=ctk.CTkFont(size=11, weight="bold"),  # Reduzir de 14 para 11
            height=30,  # Reduzir de 38 para 30
            corner_radius=6,  # Reduzir de 8 para 6
            fg_color=self.cores['accent'],
            hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary'],
            command=self.selecionar_arquivo
        )
        btn_selecionar.pack(fill="x", pady=(6, 0))  # Reduzir espaçamento
    
    def criar_selecao_idiomas(self, parent):
        """Cria a seleção de idiomas com layout organizado - versão compacta"""
        # Frame para idiomas com bordas arredondadas - mais compacto
        idiomas_frame = ctk.CTkFrame(parent, corner_radius=8, fg_color=self.cores['semi_transparent'])
        idiomas_frame.pack(fill="x", pady=(0, 10))  # Reduzir espaçamento
        
        # Título do frame - mais compacto
        ctk.CTkLabel(
            idiomas_frame,
            text="🌍 Idiomas",
            font=ctk.CTkFont(size=11, weight="bold"),  # Reduzir de 14 para 11
            text_color=self.cores['text_primary']
        ).pack(pady=8)  # Reduzir de 12 para 8
        
        # Container para os idiomas - mais compacto
        idiomas_container = ctk.CTkFrame(idiomas_frame, fg_color="transparent")
        idiomas_container.pack(fill="x", padx=10, pady=(0, 10))  # Reduzir padding
        
        # Idioma de origem - mais compacto
        origem_frame = ctk.CTkFrame(idiomas_container, fg_color="transparent")
        origem_frame.pack(fill="x", pady=(0, 10))  # Reduzir espaçamento
        
        ctk.CTkLabel(
            origem_frame,
            text="Idioma de Origem:",
            font=ctk.CTkFont(size=10, weight="bold"),  # Reduzir de 14 para 10
            text_color=self.cores['text_primary']
        ).pack(anchor="w", pady=(0, 4))  # Reduzir espaçamento
        
        idiomas = {
            'en': '🇺🇸 Inglês', 'pt': '🇧🇷 Português', 'es': '🇪🇸 Espanhol',
            'fr': '🇫🇷 Francês', 'de': '🇩🇪 Alemão', 'it': '🇮🇹 Italiano',
            'ja': '🇯🇵 Japonês', 'ko': '🇰🇷 Coreano', 'zh': '🇨🇳 Chinês',
            'ar': '🇸🇦 Árabe', 'hi': '🇮🇳 Hindi', 'ru': '🇷🇺 Russo'
        }
        
        self.combo_idioma_origem = ctk.CTkComboBox(
            origem_frame,
            values=list(idiomas.values()),
            font=ctk.CTkFont(size=10),  # Reduzir de 13 para 10
            height=26,  # Reduzir de 32 para 26
            corner_radius=5,  # Reduzir de 6 para 5
            fg_color=self.cores['glass'],
            button_color=self.cores['accent'],
            button_hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary']
        )
        self.combo_idioma_origem.set(idiomas['en'])
        self.combo_idioma_origem.pack(fill="x")
        self.combo_idioma_origem.bind("<<ComboboxSelected>>", self.atualizar_idioma_origem)
        
        # Idioma de destino - mais compacto
        destino_frame = ctk.CTkFrame(idiomas_container, fg_color="transparent")
        destino_frame.pack(fill="x")
        
        ctk.CTkLabel(
            destino_frame,
            text="Idioma de Destino:",
            font=ctk.CTkFont(size=10, weight="bold"),  # Reduzir de 14 para 10
            text_color=self.cores['text_primary']
        ).pack(anchor="w", pady=(0, 4))  # Reduzir espaçamento
        
        self.combo_idioma_destino = ctk.CTkComboBox(
            destino_frame,
            values=list(idiomas.values()),
            font=ctk.CTkFont(size=10),  # Reduzir de 13 para 10
            height=26,  # Reduzir de 32 para 26
            corner_radius=5,  # Reduzir de 6 para 5
            fg_color=self.cores['glass'],
            button_color=self.cores['accent'],
            button_hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary']
        )
        self.combo_idioma_destino.set(idiomas['pt'])
        self.combo_idioma_destino.pack(fill="x")
        self.combo_idioma_destino.bind("<<ComboboxSelected>>", self.atualizar_idioma_destino)
    
    def criar_configuracoes_avancadas(self, parent):
        """Cria as configurações avançadas com sliders organizados - versão compacta"""
        # Frame para configurações avançadas - mais compacto
        config_frame = ctk.CTkFrame(parent, corner_radius=8, fg_color=self.cores['semi_transparent'])
        config_frame.pack(fill="x", pady=(0, 10))  # Reduzir espaçamento
        
        # Título do frame - mais compacto
        ctk.CTkLabel(
            config_frame,
            text="🔧 Configurações Avançadas",
            font=ctk.CTkFont(size=11, weight="bold"),  # Reduzir de 14 para 11
            text_color=self.cores['text_primary']
        ).pack(pady=8)  # Reduzir de 12 para 8
        
        # Container para configurações - mais compacto
        config_container = ctk.CTkFrame(config_frame, fg_color="transparent")
        config_container.pack(fill="x", padx=10, pady=(0, 10))  # Reduzir padding
        
        # Tamanho do lote - mais compacto
        lote_frame = ctk.CTkFrame(config_container, fg_color="transparent")
        lote_frame.pack(fill="x", pady=(0, 10))  # Reduzir espaçamento
        
        # Label e valor em linha - mais compacto
        lote_header = ctk.CTkFrame(lote_frame, fg_color="transparent")
        lote_header.pack(fill="x")
        
        ctk.CTkLabel(
            lote_header,
            text="Tamanho do Lote:",
            font=ctk.CTkFont(size=10),  # Reduzir de 14 para 10
            text_color=self.cores['text_primary']
        ).pack(side="left")
        
        self.label_lote = ctk.CTkLabel(
            lote_header,
            text="15",
            font=ctk.CTkFont(size=10, weight="bold"),  # Reduzir de 13 para 10
            text_color=self.cores['accent']
        )
        self.label_lote.pack(side="right")
        
        # Slider com melhor aparência - mais compacto
        self.slider_lote = ctk.CTkSlider(
            lote_frame,
            from_=5,
            to=50,
            number_of_steps=45,
            height=14,  # Reduzir de 18 para 14
            corner_radius=6,  # Reduzir de 8 para 6
            fg_color=self.cores['glass'],
            progress_color=self.cores['accent'],
            button_color=self.cores['accent'],
            button_hover_color=self.cores['accent_hover']
        )
        self.slider_lote.set(15)
        self.slider_lote.pack(fill="x", pady=(6, 0))  # Reduzir espaçamento
        self.slider_lote.configure(command=self.atualizar_label_lote)
        
        # Delay - mais compacto
        delay_frame = ctk.CTkFrame(config_container, fg_color="transparent")
        delay_frame.pack(fill="x")
        
        # Label e valor em linha - mais compacto
        delay_header = ctk.CTkFrame(delay_frame, fg_color="transparent")
        delay_header.pack(fill="x")
        
        ctk.CTkLabel(
            delay_frame,
            text="Delay (ms):",
            font=ctk.CTkFont(size=10)  # Reduzir de 14 para 10
        ).pack(anchor="w", pady=(10, 3))  # Reduzir espaçamento
        
        self.label_delay = ctk.CTkLabel(
            delay_header,
            text="0.3s",
            font=ctk.CTkFont(size=10, weight="bold"),  # Reduzir de 13 para 10
            text_color=self.cores['accent']
        )
        self.label_delay.pack(side="right")
        
        # Slider com melhor aparência - mais compacto
        self.slider_delay = ctk.CTkSlider(
            delay_frame,
            from_=100,
            to=1000,
            number_of_steps=18,
            height=14,  # Reduzir de 18 para 14
            corner_radius=6,  # Reduzir de 8 para 6
            fg_color=self.cores['glass'],
            progress_color=self.cores['accent'],
            button_color=self.cores['accent'],
            button_hover_color=self.cores['accent_hover']
        )
        self.slider_delay.set(300)
        self.slider_delay.pack(fill="x", pady=(6, 0))  # Reduzir espaçamento
        self.slider_delay.configure(command=self.atualizar_label_delay)
    
    def criar_botoes_acao(self, parent):
        """Cria os botões de ação com layout organizado - versão compacta"""
        # Frame para botões - mais compacto
        botoes_frame = ctk.CTkFrame(parent, fg_color="transparent")
        botoes_frame.pack(fill="x", pady=(0, 10))  # Reduzir espaçamento
        
        # Botão salvar configurações - mais compacto
        btn_salvar = ctk.CTkButton(
            botoes_frame,
            text="Salvar Configurações",
            font=ctk.CTkFont(size=9),  # Reduzir de 11 para 9
            height=28,  # Reduzir de 36 para 28
            corner_radius=5,  # Reduzir de 6 para 5
            fg_color=self.cores['glass'],
            hover_color=self.cores['secondary_hover'],
            text_color=self.cores['text_primary'],
            command=self.salvar_configuracoes
        )
        btn_salvar.pack(fill="x")
    
    def criar_area_central(self):
        """Cria a área central com seleção de colunas e preview dos dados"""
        central_frame = ctk.CTkFrame(self.root, fg_color="transparent")
        central_frame.grid(row=0, column=1, sticky="nsew", padx=6, pady=(8, 12))  # Aproveitar espaço do header removido
        
        # Configurar grid da área central - altura equilibrada para 18 linhas
        central_frame.grid_columnconfigure(0, weight=1)
        central_frame.grid_rowconfigure(0, weight=0, minsize=80)  # Card superior ultra-compacto (altura fixa)
        central_frame.grid_rowconfigure(1, weight=1, minsize=470)  # Card inferior proporcional para 18 linhas
        
        # Card superior - Seleção de colunas (ultra-compacto)
        self.criar_card_selecao_colunas(central_frame)
        
        # Card inferior - Preview dos dados (mais espaço)
        self.criar_card_preview_dados(central_frame)
    
    def criar_card_selecao_colunas(self, parent):
        """Cria o card superior para seleção de colunas para traduzir - versão compacta"""
        # Frame principal do card - altura reduzida
        card_frame = ctk.CTkFrame(parent, corner_radius=12, fg_color=self.cores['glass'])
        card_frame.grid(row=0, column=0, sticky="ew", pady=(0, 6), padx=0)  # Espaçamento mínimo
        
        # Título do card - mais compacto
        ctk.CTkLabel(
            card_frame,
            text="🎯 Seleção de Colunas para Traduzir",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=self.cores['text_primary']
        ).pack(pady=(10, 8))
        
        # Container para as colunas - espaçamento reduzido
        colunas_container = ctk.CTkFrame(card_frame, fg_color="transparent")
        colunas_container.pack(fill="x", padx=15, pady=(0, 12))
        
        # Label explicativo - mais compacto
        ctk.CTkLabel(
            colunas_container,
            text="Selecione as colunas que deseja traduzir:",
            font=ctk.CTkFont(size=11),
            text_color=self.cores['text_secondary']
        ).pack(anchor="w", pady=(0, 6))
        
        # Frame para organizar colunas em linhas horizontais - altura reduzida
        self.colunas_container_horizontal = ctk.CTkFrame(colunas_container, fg_color="transparent")
        self.colunas_container_horizontal.pack(fill="x")
        
        # Dicionário para armazenar os checkboxes das colunas
        self.checkboxes_colunas = {}
        
        # Placeholder para as colunas (será preenchido quando carregar dataset)
        self.label_colunas_placeholder = ctk.CTkLabel(
            self.colunas_container_horizontal,
            text="📁 Carregue um dataset para ver as colunas disponíveis",
            font=ctk.CTkFont(size=11),
            text_color=self.cores['text_muted']
        )
        self.label_colunas_placeholder.pack(pady=12)
        
        # Frame para botões de tradução - mais compacto
        btn_frame = ctk.CTkFrame(colunas_container, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(8, 0))
        
        # Botão iniciar tradução - mais compacto
        self.btn_iniciar = ctk.CTkButton(
            btn_frame,
            text="🚀 Iniciar Tradução",
            font=ctk.CTkFont(size=12, weight="bold"),
            height=32,
            corner_radius=6,
            fg_color=self.cores['accent'],
            hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary'],
            command=self.iniciar_traducao
        )
        self.btn_iniciar.pack(side="left", padx=(0, 6), fill="x", expand=True)
        
        # Botão parar tradução - mais compacto
        self.btn_parar = ctk.CTkButton(
            btn_frame,
            text="⏹️ Parar Tradução",
            font=ctk.CTkFont(size=12, weight="bold"),
            height=32,
            corner_radius=6,
            fg_color=self.cores['warning'],
            hover_color=self.cores['warning_hover'],
            text_color=self.cores['text_primary'],
            command=self.parar_traducao,
            state="disabled"
        )
        self.btn_parar.pack(side="left", padx=(6, 0), fill="x", expand=True)
    
    def criar_card_preview_dados(self, parent):
        """Cria o card inferior para preview dos dados"""
        # Frame principal do card
        card_frame = ctk.CTkFrame(parent, corner_radius=15, fg_color=self.cores['secondary'])
        card_frame.grid(row=1, column=0, sticky="nsew", pady=(2, 0), padx=0)  # Espaçamento mínimo
        
        # Configurar grid do card - altura proporcional para 18 linhas
        card_frame.grid_columnconfigure(0, weight=1)
        card_frame.grid_rowconfigure(0, weight=0, minsize=35)  # Header ultra-compacto
        card_frame.grid_rowconfigure(1, weight=1, minsize=400)  # Área de dados proporcional para 18 linhas
        
        # Header do card com título e controles - apenas informações do dataset
        header_frame = ctk.CTkFrame(card_frame, fg_color="transparent")
        header_frame.grid(row=0, column=0, sticky="ew", padx=12, pady=4)  # Header ultra-compacto
        
        # Título e informações - mais compacto
        info_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        info_frame.pack(side="left", fill="y")
        
        ctk.CTkLabel(
            info_frame,
            text="📊 Preview dos Dados",
            font=ctk.CTkFont(size=13, weight="bold"),  # Fonte menor
            text_color=self.cores['text_primary']
        ).pack(anchor="w")
        
        # Label para informações do dataset - mais compacto
        self.label_info_dataset = ctk.CTkLabel(
            info_frame,
            text="Nenhum dataset carregado",
            font=ctk.CTkFont(size=10),  # Fonte menor
            text_color=self.cores['text_muted']
        )
        self.label_info_dataset.pack(anchor="w", pady=(2, 0))  # Espaçamento reduzido
        
        # Remover controles da direita - apenas informações do dataset
        
        # Área de visualização dos dados - espaçamento equilibrado
        data_frame = ctk.CTkFrame(card_frame, fg_color="transparent")
        data_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(3, 10))  # Espaçamento equilibrado para 18 linhas
        
        # Frame para a tabela
        table_frame = ctk.CTkFrame(data_frame, fg_color="transparent")
        table_frame.pack(fill="both", expand=True)
        
        # Criar Treeview para mostrar os dados como tabela
        import tkinter as tk
        from tkinter import ttk
        
        # Frame para conter o Treeview
        tree_container = ctk.CTkFrame(table_frame, fg_color="transparent")
        tree_container.pack(fill="both", expand=True)
        
        # Criar Treeview
        self.tree = ttk.Treeview(tree_container, show="headings", height=20)  # Altura proporcional para 18 linhas
        
        # Configurar scrollbars personalizadas
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Layout da tabela
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Configurar grid weights
        tree_container.grid_columnconfigure(0, weight=1)
        tree_container.grid_rowconfigure(0, weight=1)
        
        # Configurar estilo da tabela para tema escuro
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Treeview",
            background=self.cores['glass'],
            foreground=self.cores['text_primary'],
            fieldbackground=self.cores['glass'],
            borderwidth=0,
            font=('Consolas', 9)
        )
        style.configure(
            "Treeview.Heading",
            background=self.cores['accent'],
            foreground=self.cores['text_primary'],
            font=('Consolas', 9, 'bold')
        )
        style.map(
            "Treeview",
            background=[('selected', self.cores['accent'])],
            foreground=[('selected', self.cores['text_primary'])]
        )
        
        # Personalizar barras de rolagem para combinar com o tema
        style.configure(
            "Vertical.TScrollbar",
            background=self.cores['glass'],
            troughcolor=self.cores['secondary'],
            bordercolor=self.cores['border'],
            arrowcolor=self.cores['text_secondary'],
            width=12  # Largura da barra vertical
        )
        style.configure(
            "Horizontal.TScrollbar",
            background=self.cores['glass'],
            troughcolor=self.cores['secondary'],
            bordercolor=self.cores['border'],
            arrowcolor=self.cores['text_secondary'],
            height=12  # Altura da barra horizontal
        )
        
        # Aplicar estilos personalizados às barras de rolagem
        vsb.configure(style="Vertical.TScrollbar")
        hsb.configure(style="Horizontal.TScrollbar")
        
        # Adicionar efeitos hover para as barras de rolagem
        style.map(
            "Vertical.TScrollbar",
            background=[('active', self.cores['accent']), ('pressed', self.cores['accent_hover'])],
            troughcolor=[('active', self.cores['secondary_hover'])]
        )
        style.map(
            "Horizontal.TScrollbar",
            background=[('active', self.cores['accent']), ('pressed', self.cores['accent_hover'])],
            troughcolor=[('active', self.cores['secondary_hover'])]
        )
        

    
    def criar_sidebar_direita(self):
        """Cria a sidebar direita focada nos logs de atividades - versão compacta"""
        sidebar_frame = ctk.CTkFrame(self.root, corner_radius=10, fg_color=self.cores['glass'])
        sidebar_frame.grid(row=0, column=2, sticky="nsew", padx=(6, 12), pady=(8, 12))  # Aproveitar espaço do header removido
        
        # Título da sidebar - mais compacto
        sidebar_title = ctk.CTkLabel(
            sidebar_frame,
            text="📝 Log de Atividades",
            font=ctk.CTkFont(size=13, weight="bold"),  # Reduzir de 16 para 13
            text_color=self.cores['text_primary']
        )
        sidebar_title.pack(pady=8)  # Reduzir de 12 para 8
        
        # Container para logs - mais compacto
        log_container = ctk.CTkScrollableFrame(
            sidebar_frame, 
            fg_color="transparent"
        )
        log_container.pack(fill="both", expand=True, padx=12, pady=(0, 12))  # Reduzir padding
        
        # Log de atividades
        self.criar_log_atividades(log_container)
    
    def criar_log_atividades(self, parent):
        """Cria o log de atividades com layout organizado - versão compacta"""
        log_frame = ctk.CTkFrame(parent, corner_radius=8, fg_color=self.cores['semi_transparent'])
        log_frame.pack(fill="both", expand=True)
        
        # Título do frame - mais compacto
        ctk.CTkLabel(
            log_frame,
            text="📋 Histórico de Atividades",
            font=ctk.CTkFont(size=11, weight="bold"),  # Reduzir de 14 para 11
            text_color=self.cores['text_primary']
        ).pack(pady=8)  # Reduzir de 12 para 8
        
        # Textbox para log - altura proporcional
        self.textbox_log = ctk.CTkTextbox(
            log_frame,
            font=ctk.CTkFont(size=9, family="Consolas"),  # Reduzir de 11 para 9
            corner_radius=8,  # Reduzir de 10 para 8
            height=350  # Aumentar para 350 para ficar proporcional
        )
        self.textbox_log.pack(fill="both", expand=True, padx=12, pady=(0, 12))  # Reduzir padding
        
        # Botão limpar log (centralizado) - mais compacto
        btn_limpar = ctk.CTkButton(
            log_frame,
            text="Limpar Log",
            font=ctk.CTkFont(size=9),  # Reduzir de 11 para 9
            height=28,  # Reduzir de 36 para 28
            corner_radius=5,  # Reduzir de 6 para 5
            fg_color=self.cores['glass'],
            hover_color=self.cores['secondary_hover'],
            text_color=self.cores['text_primary'],
            command=self.limpar_log
        )
        btn_limpar.pack(pady=(0, 12))  # Reduzir espaçamento
        
        # Separador visual
        separator = ctk.CTkFrame(log_frame, height=1, fg_color=self.cores['border'])
        separator.pack(fill="x", pady=(0, 8))
        
        # Botões de exportação - mais compactos
        export_label = ctk.CTkLabel(
            log_frame,
            text="💾 Exportar Dataset:",
            font=ctk.CTkFont(size=9, weight="bold"),
            text_color=self.cores['text_primary']
        )
        export_label.pack(anchor="w", pady=(0, 6))
        
        # Botão exportar CSV
        btn_csv = ctk.CTkButton(
            log_frame,
            text="📊 Exportar CSV",
            font=ctk.CTkFont(size=9),
            height=26,
            corner_radius=5,
            fg_color=self.cores['accent'],
            hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary'],
            command=lambda: self.exportar_resultado('CSV')
        )
        btn_csv.pack(fill="x", pady=(0, 4))
        
        # Botão exportar Excel
        btn_excel = ctk.CTkButton(
            log_frame,
            text="📈 Exportar Excel",
            font=ctk.CTkFont(size=9),
            height=26,
            corner_radius=5,
            fg_color=self.cores['accent'],
            hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary'],
            command=lambda: self.exportar_resultado('Excel')
        )
        btn_excel.pack(fill="x", pady=(0, 0))
    
    def criar_barra_status(self):
        """Cria a barra de status inferior organizada e responsiva - versão compacta com progresso"""
        status_bar = ctk.CTkFrame(self.root, corner_radius=0, fg_color=self.cores['glass'])
        status_bar.grid(row=1, column=0, columnspan=3, sticky="ew", padx=0, pady=(0, 0))  # Ajustar para nova estrutura sem header
        
        # Status à esquerda - mais compacto
        self.label_status_bar = ctk.CTkLabel(
            status_bar,
            text="Pronto",
            font=ctk.CTkFont(size=10),  # Reduzir de 12 para 10
            text_color=self.cores['text_secondary']
        )
        self.label_status_bar.pack(side="left", padx=12, pady=6)  # Reduzir padding
        
        # Barra de progresso no centro - mais compacta
        progress_frame = ctk.CTkFrame(status_bar, fg_color="transparent")
        progress_frame.pack(side="left", padx=20, pady=6)
        
        ctk.CTkLabel(
            progress_frame,
            text="Progresso:",
            font=ctk.CTkFont(size=9, weight="bold"),
            text_color=self.cores['text_primary']
        ).pack(side="left", padx=(0, 5))
        
        self.progress_bar = ctk.CTkProgressBar(
            progress_frame,
            height=10,  # Altura reduzida
            width=100,  # Largura fixa para não ocupar muito espaço
            corner_radius=5,
            fg_color=self.cores['glass'],
            progress_color=self.cores['accent']
        )
        self.progress_bar.pack(side="left", padx=(0, 5))
        self.progress_bar.set(0)
        
        # Label para mostrar porcentagem - mais compacto
        self.label_progress = ctk.CTkLabel(
            progress_frame,
            text="0%",
            font=ctk.CTkFont(size=9),
            text_color=self.cores['accent']
        )
        self.label_progress.pack(side="left")
        
        # Status da tradução à direita - mais compacto
        self.label_status_traducao = ctk.CTkLabel(
            status_bar,
            text="Pronto para traduzir",
            font=ctk.CTkFont(size=9),
            text_color=self.cores['text_secondary']
        )
        self.label_status_traducao.pack(side="right", padx=12, pady=6)
        
        # Informações à direita - mais compacto
        self.label_info = ctk.CTkLabel(
            status_bar,
            text="",
            font=ctk.CTkFont(size=9),  # Reduzir de 10 para 9
            text_color=self.cores['text_muted']
        )
        self.label_info.pack(side="right", padx=12, pady=6)  # Reduzir padding
    
    # Métodos de funcionalidade (implementação básica)
    def atualizar_visibilidade_tabela_sqlite(self):
        """Atualiza a visibilidade do frame de tabela SQLite baseado no tipo de arquivo"""
        if hasattr(self, 'tipo_arquivo') and self.tipo_arquivo.get() == "SQLite":
            if hasattr(self, 'label_tabela_placeholder'):
                self.label_tabela_placeholder.pack_forget()
        else:
            if hasattr(self, 'frame_tabela_sqlite'):
                self.frame_tabela_sqlite.pack_forget()
            if hasattr(self, 'label_tabela_placeholder'):
                self.label_tabela_placeholder.pack(pady=20)
    
    def feedback_visual(self, widget, cor_original=None, duracao=300):
        """Cria um efeito de feedback visual temporário no widget"""
        try:
            if cor_original is None:
                cor_original = widget.cget("fg_color")
            
            # Cor de destaque
            cor_destaque = self.cores['accent']
            
            # Aplicar cor de destaque
            widget.configure(fg_color=cor_destaque)
            
            # Restaurar cor original após o tempo especificado
            self.root.after(duracao, lambda: widget.configure(fg_color=cor_original))
            
        except Exception:
            pass  # Ignorar erros de feedback visual
    
    def trocar_tabela_sqlite(self, event=None):
        """Permite trocar de tabela no banco SQLite"""
        try:
            if hasattr(self, 'df_full_path') and self.df_full_path and self.df_tipo == "SQLite":
                nova_tabela = self.combo_tabela_sqlite.get()
                if nova_tabela and nova_tabela != "Nenhuma tabela selecionada":
                    self.df_tabela = nova_tabela
                    self.carregar_arquivo(self.df_full_path, "SQLite")
                    self.log_atividade(f"Tabela SQLite alterada para: {nova_tabela}")
        except Exception as exc:
            self.log_atividade(f"ERRO ao trocar tabela SQLite: {exc}")
            self.mostrar_dialogo_personalizado("Erro", f"Falha ao trocar tabela:\n\n{exc}", "error")
    
    def atualizar_label_lote(self, value):
        """Atualiza o label do slider de lote"""
        self.label_lote.configure(text=str(int(value)))
    
    def atualizar_label_delay(self, value):
        """Atualiza o label do slider de delay (converte para segundos)"""
        delay_ms = int(value)
        delay_s = delay_ms / 1000.0
        self.label_delay.configure(text=f"{delay_s:.1f}s")
    
    def atualizar_idioma_origem(self, event=None):
        """Atualiza o idioma de origem"""
        idiomas = {
            '🇺🇸 Inglês': 'en', '🇧🇷 Português': 'pt', '🇪🇸 Espanhol': 'es',
            '🇫🇷 Francês': 'fr', '🇩🇪 Alemão': 'de', '🇮🇹 Italiano': 'it',
            '🇯🇵 Japonês': 'ja', '🇰🇷 Coreano': 'ko', '🇨🇳 Chinês': 'zh',
            '🇸🇦 Árabe': 'ar', '🇮🇳 Hindi': 'hi', '🇷🇺 Russo': 'ru'
        }
        self.config['idioma_origem'] = idiomas[self.combo_idioma_origem.get()]
        self.log_atividade(f"Idioma de origem alterado para: {self.combo_idioma_origem.get()}")
    
    def atualizar_idioma_destino(self, event=None):
        """Atualiza o idioma de destino"""
        idiomas = {
            '🇺🇸 Inglês': 'en', '🇧🇷 Português': 'pt', '🇪🇸 Espanhol': 'es',
            '🇫🇷 Francês': 'fr', '🇩🇪 Alemão': 'de', '🇮🇹 Italiano': 'it',
            '🇯🇵 Japonês': 'ja', '🇰🇷 Coreano': 'ko', '🇨🇳 Chinês': 'zh',
            '🇸🇦 Árabe': 'ar', '🇮🇳 Hindi': 'hi', '🇷🇺 Russo': 'ru'
        }
        self.config['idioma_destino'] = idiomas[self.combo_idioma_destino.get()]
        self.log_atividade(f"Idioma de destino alterado para: {self.combo_idioma_destino.get()}")
    
    def selecionar_arquivo(self):
        """Abre diálogo para selecionar arquivo"""
        try:
            tipo = self.tipo_arquivo.get() if hasattr(self, "tipo_arquivo") else "CSV"
            filename = None
            if tipo == "CSV":
                filename = filedialog.askopenfilename(
                    title="Selecionar arquivo CSV",
                    filetypes=[("CSV", "*.csv"), ("Todos", "*.*")]
                )
            elif tipo == "Excel":
                filename = filedialog.askopenfilename(
                    title="Selecionar arquivo Excel",
                    filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
                )
            elif tipo == "SQLite":
                filename = filedialog.askopenfilename(
                    title="Selecionar banco SQLite",
                    filetypes=[("SQLite", "*.db *.sqlite *.sqlite3"), ("Todos", "*.*")]
                )

            if filename:
                # Feedback visual no botão
                for widget in self.root.winfo_children():
                    if isinstance(widget, ctk.CTkFrame):
                        for child in widget.winfo_children():
                            if isinstance(child, ctk.CTkScrollableFrame):
                                for grandchild in child.winfo_children():
                                    if isinstance(grandchild, ctk.CTkFrame):
                                        for button in grandchild.winfo_children():
                                            if isinstance(button, ctk.CTkButton) and button.cget("text") == "Selecionar Arquivo":
                                                self.feedback_visual(button)
                                                break
                
                # Atualizar visibilidade da tabela SQLite
                self.atualizar_visibilidade_tabela_sqlite()
                
                self.carregar_arquivo(filename, tipo)
        except Exception as exc:
            self.mostrar_dialogo_personalizado("Erro", f"Falha ao selecionar arquivo:\n\n{exc}", "error")
            self.log_atividade(f"ERRO ao selecionar arquivo: {exc}")

    def obter_tabela_sqlite(self, filename: str):
        """Pergunta ao usuário qual tabela abrir, caso haja mais de uma."""
        try:
            conn = sqlite3.connect(filename)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tabelas = [row[0] for row in cursor.fetchall()]
            conn.close()

            if not tabelas:
                self.mostrar_dialogo_personalizado("Aviso", "Nenhuma tabela encontrada no banco SQLite", "warning")
                return None

            if len(tabelas) == 1:
                return tabelas[0]

            # Diálogo melhorado usando CTkToplevel
            dialog = ctk.CTkToplevel(self.root)
            dialog.title("🗄️ Selecionar Tabela SQLite")
            dialog.geometry("400x200")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.resizable(False, False)
            
            # Centralizar diálogo
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
            y = (dialog.winfo_screenheight() // 2) - (200 // 2)
            dialog.geometry(f"400x200+{x}+{y}")

            # Título
            ctk.CTkLabel(
                dialog, 
                text="🗄️ Selecionar Tabela", 
                font=ctk.CTkFont(size=16, weight="bold")
            ).pack(pady=(20, 10))
            
            # Descrição
            ctk.CTkLabel(
                dialog, 
                text=f"Banco: {os.path.basename(filename)}\nTabelas encontradas: {len(tabelas)}",
                font=ctk.CTkFont(size=12),
                text_color="gray"
            ).pack(pady=(0, 15))

            # Combo box para seleção
            combo = ctk.CTkComboBox(
                dialog, 
                values=tabelas, 
                width=300,
                fg_color=self.cores['glass'],
                button_color=self.cores['accent'],
                button_hover_color=self.cores['accent_hover']
            )
            combo.set(tabelas[0])
            combo.pack(pady=(0, 20))

            escolha = {"tabela": None}

            def confirmar():
                escolha["tabela"] = combo.get()
                dialog.destroy()

            # Botão confirmar
            ctk.CTkButton(
                dialog, 
                text="Confirmar", 
                command=confirmar,
                width=120,
                fg_color=self.cores['accent'],
                hover_color=self.cores['accent_hover']
            ).pack()
            
            dialog.wait_window()
            return escolha["tabela"]
        except Exception as exc:
            messagebox.showerror("Erro", f"Falha ao listar tabelas: {exc}")
            return None

    def carregar_arquivo(self, filename: str, tipo: str):
        """Carrega o arquivo selecionado APENAS para preview - SEM traduzir"""
        try:
            if tipo == "CSV":
                # Para CSV, ler apenas as primeiras linhas para preview
                self.df_preview = pd.read_csv(filename, nrows=18)
                self.df_full_path = filename
                self.df_tipo = "CSV"
                self.df_tabela = None
            elif tipo == "Excel":
                # Para Excel, ler apenas as primeiras linhas para preview
                self.df_preview = pd.read_excel(filename, nrows=18)
                self.df_full_path = filename
                self.df_tipo = "Excel"
                self.df_tabela = None
            elif tipo == "SQLite":
                # Para SQLite, selecionar tabela e ler preview
                tabela = self.obter_tabela_sqlite(filename)
                if not tabela:
                    return
                conn = sqlite3.connect(filename)
                try:
                    # Ler apenas as primeiras linhas para preview
                    self.df_preview = pd.read_sql_query(f"SELECT * FROM {tabela} LIMIT 18", conn)
                    self.df_full_path = filename
                    self.df_tipo = "SQLite"
                    self.df_tabela = tabela
                    
                    # Atualizar combo de tabelas disponíveis
                    self.atualizar_combo_tabelas_sqlite(filename)
                finally:
                    conn.close()
            else:
                messagebox.showwarning("Aviso", f"Tipo de arquivo não suportado: {tipo}")
                return

            # Atualizar interface com preview
            self.colunas_originais = list(self.df_preview.columns)
            self.atualizar_interface_dataset(filename)
            self.atualizar_previa_dados()
            
            # Log da atividade
            self.log_atividade(f"Dataset carregado (preview): {os.path.basename(filename)} | Linhas no preview: {len(self.df_preview)}, Colunas: {len(self.colunas_originais)}")
            
            # Mostrar mensagem informativa personalizada
            if tipo == "SQLite":
                self.mostrar_dialogo_personalizado(
                    "Sucesso", 
                    f"Dataset carregado com sucesso!\n\nTabela: {self.df_tabela}\nLinhas no preview: {len(self.df_preview)}\nColunas: {len(self.colunas_originais)}\n\n✅ Preview carregado. Agora selecione as colunas para traduzir e clique em 'Iniciar Tradução'.",
                    "info"
                )
            else:
                self.mostrar_dialogo_personalizado(
                    "Sucesso", 
                    f"Dataset carregado com sucesso!\n\nLinhas no preview: {len(self.df_preview)}\nColunas: {len(self.colunas_originais)}\n\n⚠️ Apenas as primeiras 18 linhas foram carregadas para preview.\nO arquivo completo será processado durante a tradução.",
                    "warning"
                )
                
        except Exception as exc:
            self.mostrar_dialogo_personalizado("Erro", f"Erro ao carregar arquivo:\n\n{exc}", "error")
            self.log_atividade(f"ERRO ao carregar arquivo: {exc}")

    def atualizar_combo_tabelas_sqlite(self, filename: str):
        """Atualiza o combo de tabelas disponíveis no banco SQLite"""
        try:
            conn = sqlite3.connect(filename)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tabelas = [row[0] for row in cursor.fetchall()]
            conn.close()
            
            if tabelas:
                # Atualizar valores do combo
                self.combo_tabela_sqlite.configure(values=tabelas)
                # Manter a tabela atual selecionada
                if self.df_tabela in tabelas:
                    self.combo_tabela_sqlite.set(self.df_tabela)
                else:
                    self.combo_tabela_sqlite.set(tabelas[0])
                
                # Mostrar o frame de tabela e ocultar placeholder
                self.frame_tabela_sqlite.pack(fill="x", pady=(0, 15))
                if hasattr(self, 'label_tabela_placeholder'):
                    self.label_tabela_placeholder.pack_forget()
            else:
                self.combo_tabela_sqlite.configure(values=["Nenhuma tabela encontrada"])
                self.combo_tabela_sqlite.set("Nenhuma tabela encontrada")
                # Ocultar o frame de tabela e mostrar placeholder
                self.frame_tabela_sqlite.pack_forget()
                if hasattr(self, 'label_tabela_placeholder'):
                    self.label_tabela_placeholder.pack(pady=20)
                
        except Exception as exc:
            self.log_atividade(f"ERRO ao atualizar combo de tabelas SQLite: {exc}")
    
    def atualizar_interface_dataset(self, filename: str = ""):
        """Atualiza labels, lista de colunas e estatísticas com base no preview."""
        if self.df_preview is None:
            return
            
        # Criar checkboxes para seleção de colunas
        self.criar_checkboxes_colunas()
        
        # Atualizar informações do dataset no card de preview
        base = os.path.basename(filename) if filename else ""
        tipo_info = f" ({self.df_tipo})" if hasattr(self, 'df_tipo') else ""
        tabela_info = f" - Tabela: {self.df_tabela}" if hasattr(self, 'df_tipo') and self.df_tipo == "SQLite" and hasattr(self, 'df_tabela') and self.df_tabela else ""
        
        info_txt = f"📁 {base}{tipo_info}{tabela_info} | 📊 {len(self.df_preview):,} linhas | 🗂️ {len(self.colunas_originais)} colunas"
        self.label_info_dataset.configure(text=info_txt)
        
        # Log da atividade
        self.log_atividade(f"Dataset carregado: {base}{tipo_info}{tabela_info} | Preview: {len(self.df_preview):,} linhas, {len(self.colunas_originais)} colunas")

    def criar_checkboxes_colunas(self):
        """Cria checkboxes para seleção de colunas para traduzir organizados horizontalmente - versão compacta"""
        # Limpar checkboxes existentes
        for checkbox in self.checkboxes_colunas.values():
            checkbox.destroy()
        self.checkboxes_colunas.clear()
        
        # Remover placeholder se existir
        if hasattr(self, 'label_colunas_placeholder'):
            self.label_colunas_placeholder.destroy()
            delattr(self, 'label_colunas_placeholder')
        
        # Calcular quantas colunas cabem por linha (mais colunas por linha para economizar espaço)
        colunas_por_linha = 6  # Aumentar de 4 para 6
        total_linhas = (len(self.colunas_originais) + colunas_por_linha - 1) // colunas_por_linha
        
        # Criar linhas horizontais de checkboxes
        for linha in range(total_linhas):
            # Frame para cada linha horizontal - mais compacto
            linha_frame = ctk.CTkFrame(self.colunas_container_horizontal, fg_color="transparent")
            linha_frame.pack(fill="x", pady=1)  # Reduzir espaçamento entre linhas
            
            # Configurar grid da linha para distribuir os checkboxes uniformemente
            for i in range(colunas_por_linha):
                linha_frame.grid_columnconfigure(i, weight=1)
            
            # Adicionar checkboxes nesta linha
            for col_idx in range(colunas_por_linha):
                col_global = linha * colunas_por_linha + col_idx
                
                if col_global < len(self.colunas_originais):
                    col = self.colunas_originais[col_global]
                    
                    # Checkbox para a coluna - mais compacto
                    checkbox = ctk.CTkCheckBox(
                        linha_frame,
                        text=col,
                        font=ctk.CTkFont(size=10),  # Fonte menor
                        text_color=self.cores['text_primary'],
                        fg_color=self.cores['accent'],
                        hover_color=self.cores['accent_hover'],
                        corner_radius=3,  # Bordas mais arredondadas
                        checkbox_width=14,  # Checkbox menor
                        checkbox_height=14
                    )
                    checkbox.grid(row=0, column=col_idx, padx=2, pady=1, sticky="ew")  # Espaçamento reduzido
                    
                    # Armazenar referência
                    self.checkboxes_colunas[col] = checkbox
                    
                    # Marcar automaticamente as primeiras colunas (opcional)
                    if col_global < 3:  # Marcar as 3 primeiras colunas por padrão
                        checkbox.select()

    def atualizar_previa_dados(self):
        """Mostra as primeiras linhas do DataFrame na área de prévia como tabela."""
        if self.df_preview is None:
            return
        try:
            # Limpar tabela existente
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Configurar colunas da tabela
            colunas = list(self.df_preview.columns)
            self.tree["columns"] = colunas
            
            # Configurar cabeçalhos das colunas
            for col in colunas:
                self.tree.heading(col, text=col)
                # Ajustar largura da coluna baseado no conteúdo
                max_width = max(len(str(col)), 10)
                for value in self.df_preview[col].head(20):
                    if pd.notna(value):
                        max_width = max(max_width, len(str(value)))
                self.tree.column(col, width=min(max_width * 8, 200), minwidth=80)
            
            # Inserir dados (primeiras 18 linhas)
            for idx, row in self.df_preview.head(18).iterrows():
                values = []
                for col in colunas:
                    value = row[col]
                    if pd.isna(value):
                        values.append("")
                    else:
                        # Truncar valores muito longos
                        str_value = str(value)
                        if len(str_value) > 50:
                            str_value = str_value[:47] + "..."
                        values.append(str_value)
                
                self.tree.insert("", "end", values=values)
            
            # Log de sucesso
            self.log_atividade(f"Preview da tabela atualizado: {len(self.df_preview.head(18))} linhas exibidas")
            
        except Exception as exc:
            self.log_atividade(f"❌ Falha ao gerar preview da tabela: {exc}")
            # Fallback para texto simples se a tabela falhar
            try:
                # Criar um textbox temporário para mostrar erro
                if not hasattr(self, 'textbox_fallback'):
                    self.textbox_fallback = ctk.CTkTextbox(
                        self.tree.master,
                        font=ctk.CTkFont(size=10, family="Consolas"),
                        corner_radius=10,
                        fg_color=self.cores['glass']
                    )
                    self.textbox_fallback.pack(fill="both", expand=True)
                
                self.textbox_fallback.delete("0.0", "end")
                self.textbox_fallback.insert("end", f"❌ Erro ao carregar tabela: {exc}\n\n")
                self.textbox_fallback.insert("end", "Dados em formato texto:\n")
                self.textbox_fallback.insert("end", "-" * 40 + "\n")
                preview = self.df_preview.to_string(index=False, max_cols=10, max_colwidth=30)
                self.textbox_fallback.insert("end", preview)
            except:
                pass
    
    def iniciar_traducao(self):
        """Inicia o processo de tradução com gerenciamento de memória inteligente"""
        if not hasattr(self, 'df_full_path') or not self.df_full_path:
            self.mostrar_dialogo_personalizado("Aviso", "Nenhum dataset carregado.\n\nCarregue um arquivo primeiro.", "warning")
            return
            
        if self.traducao_ativa:
            messagebox.showinfo("Info", "Tradução já está em andamento.")
            return
            
        # Obter colunas selecionadas dos checkboxes
        colunas_selecionadas = []
        for col, checkbox in self.checkboxes_colunas.items():
            if checkbox.get():  # Se o checkbox está marcado
                colunas_selecionadas.append(col)
        
        if not colunas_selecionadas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma coluna para traduzir.")
            return
            
        # Confirmar início da tradução
        mensagem_confirmacao = f"Iniciar tradução?\n\nArquivo: {os.path.basename(self.df_full_path)}\nTipo: {self.df_tipo}\nColunas: {len(colunas_selecionadas)}\nIdioma: {self.combo_idioma_origem.get()} → {self.combo_idioma_destino.get()}\n\n⚠️ O arquivo será processado em lotes pequenos para economizar memória.\n💡 Use o botão 'Parar Tradução' a qualquer momento para interromper."
        
        resposta = self.mostrar_confirmacao_personalizada("Confirmar Tradução", mensagem_confirmacao)
        
        if not resposta:
            return
            
        # Iniciar tradução em thread separada
        self.traducao_ativa = True
        self.thread_traducao = threading.Thread(
            target=self._executar_traducao,
            args=(colunas_selecionadas,)
        )
        self.thread_traducao.daemon = True
        self.thread_traducao.start()
        
        # Atualizar interface
        self.btn_iniciar.configure(state="disabled")
        self.btn_parar.configure(state="normal")
        self.label_status_bar.configure(text="🔄 Traduzindo...")
        self.log_atividade("Tradução iniciada")
        
        # Iniciar monitoramento de progresso
        self.monitorar_progresso()
    
    def _executar_traducao(self, colunas_selecionadas):
        """Executa a tradução em lotes para economizar memória"""
        try:
            # Configurar tradutor
            idioma_origem = self.config['idioma_origem']
            idioma_destino = self.config['idioma_destino']
            tamanho_lote = min(self.config['tamanho_lote'], 20)  # Máximo 20 linhas por lote para economizar RAM
            delay = self.config['delay_traducao']
            
            tradutor = GoogleTranslator(source=idioma_origem, target=idioma_destino)
            
            # Verificar se deve parar antes de começar
            if not self.traducao_ativa:
                return
            
            # Carregar arquivo completo em lotes
            if self.df_tipo == "CSV":
                # Para CSV, usar pandas em lotes
                self._traduzir_csv_lotes(colunas_selecionadas, tradutor, tamanho_lote, delay)
            elif self.df_tipo == "Excel":
                # Para Excel, usar pandas em lotes
                self._traduzir_excel_lotes(colunas_selecionadas, tradutor, tamanho_lote, delay)
            elif self.df_tipo == "SQLite":
                # Para SQLite, usar queries em lotes
                self._traduzir_sqlite_lotes(colunas_selecionadas, tradutor, tamanho_lote, delay)
                
        except Exception as e:
            self.progress_queue.put(("erro", f"Erro na tradução: {str(e)}"))
        finally:
            self.traducao_ativa = False
            self.progress_queue.put(("concluido", "Tradução concluída"))
    
    def _traduzir_csv_lotes(self, colunas_selecionadas, tradutor, tamanho_lote, delay):
        """Traduz CSV em lotes para economizar memória"""
        try:
            # Contar total de linhas
            total_linhas = sum(1 for _ in open(self.df_full_path, 'r', encoding='utf-8')) - 1  # -1 para header
            
            # Processar em lotes
            for i in range(0, total_linhas, tamanho_lote):
                # Verificar se deve parar a cada lote
                if not self.traducao_ativa:
                    self.log_atividade("Tradução interrompida pelo usuário")
                    break
                    
                # Ler lote
                df_lote = pd.read_csv(self.df_full_path, skiprows=i+1, nrows=tamanho_lote, header=None)
                df_lote.columns = self.colunas_originais
                
                # Traduzir colunas selecionadas
                for col in colunas_selecionadas:
                    if col in df_lote.columns:
                        df_lote[f"{col}_traduzido"] = df_lote[col].apply(
                            lambda x: tradutor.translate(str(x)) if pd.notna(x) else x
                        )
                
                # Salvar lote traduzido
                # TODO: Implementar salvamento em arquivo temporário ou banco
                
                # Atualizar progresso
                progresso = min(100, (i + tamanho_lote) / total_linhas * 100)
                self.progress_queue.put(("progresso", progresso))
                
                # Log de progresso
                self.log_atividade(f"Lote processado: {i+1}-{min(i+tamanho_lote, total_linhas)} de {total_linhas} linhas")
                
                # Delay para não sobrecarregar API
                time.sleep(delay)
                
                # Liberar memória do lote
                del df_lote
                
        except Exception as e:
            self.progress_queue.put(("erro", f"Erro ao traduzir CSV: {str(e)}"))
    
    def _traduzir_excel_lotes(self, colunas_selecionadas, tradutor, tamanho_lote, delay):
        """Traduz Excel em lotes para economizar memória"""
        try:
            # Para Excel, usar openpyxl para leitura em lotes
            from openpyxl import load_workbook
            
            wb = load_workbook(self.df_full_path, read_only=True)
            ws = wb.active
            
            # Contar linhas
            total_linhas = ws.max_row - 1  # -1 para header
            
            # Processar em lotes
            for i in range(0, total_linhas, tamanho_lote):
                # Verificar se deve parar a cada lote
                if not self.traducao_ativa:
                    self.log_atividade("Tradução interrompida pelo usuário")
                    break
                    
                # Ler lote
                dados_lote = []
                for row_idx in range(i + 2, min(i + 2 + tamanho_lote, total_linhas + 2)):  # +2 para pular header
                    row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(self.colunas_originais) + 1)]
                    dados_lote.append(row_data)
                
                df_lote = pd.DataFrame(dados_lote, columns=self.colunas_originais)
                
                # Traduzir colunas selecionadas
                for col in colunas_selecionadas:
                    if col in df_lote.columns:
                        df_lote[f"{col}_traduzido"] = df_lote[col].apply(
                            lambda x: tradutor.translate(str(x)) if pd.notna(x) else x
                        )
                
                # Atualizar progresso
                progresso = min(100, (i + tamanho_lote) / total_linhas * 100)
                self.progress_queue.put(("progresso", progresso))
                
                # Log de progresso
                self.log_atividade(f"Lote processado: {i+1}-{min(i+tamanho_lote, total_linhas)} de {total_linhas} linhas")
                
                # Delay para não sobrecarregar API
                time.sleep(delay)
                
                # Liberar memória do lote
                del df_lote, dados_lote
            
            wb.close()
            
        except Exception as e:
            self.progress_queue.put(("erro", f"Erro ao traduzir Excel: {str(e)}"))
    
    def _traduzir_sqlite_lotes(self, colunas_selecionadas, tradutor, tamanho_lote, delay):
        """Traduz SQLite em lotes para economizar memória"""
        try:
            conn = sqlite3.connect(self.df_full_path)
            
            # Contar total de linhas
            cursor = conn.cursor()
            cursor.execute(f"SELECT COUNT(*) FROM {self.df_tabela}")
            total_linhas = cursor.fetchone()[0]
            
            # Processar em lotes
            for i in range(0, total_linhas, tamanho_lote):
                # Verificar se deve parar a cada lote
                if not self.traducao_ativa:
                    self.log_atividade("Tradução interrompida pelo usuário")
                    break
                    
                # Ler lote
                query = f"SELECT * FROM {self.df_tabela} LIMIT {tamanho_lote} OFFSET {i}"
                df_lote = pd.read_sql_query(query, conn)
                
                # Traduzir colunas selecionadas
                for col in colunas_selecionadas:
                    if col in df_lote.columns:
                        df_lote[f"{col}_traduzido"] = df_lote[col].apply(
                            lambda x: tradutor.translate(str(x)) if pd.notna(x) else x
                        )
                
                # Atualizar progresso
                progresso = min(100, (i + tamanho_lote) / total_linhas * 100)
                self.progress_queue.put(("progresso", progresso))
                
                # Log de progresso
                self.log_atividade(f"Lote processado: {i+1}-{min(i+tamanho_lote, total_linhas)} de {tamanho_lote} linhas")
                
                # Delay para não sobrecarregar API
                time.sleep(delay)
                
                # Liberar memória do lote
                del df_lote
            
            conn.close()
            
        except Exception as e:
            self.progress_queue.put(("erro", f"Erro ao traduzir SQLite: {str(e)}"))
    
    def parar_traducao(self):
        """Para a tradução em andamento"""
        if not self.traducao_ativa:
            messagebox.showinfo("Info", "Nenhuma tradução em andamento.")
            return
            
        resposta = self.mostrar_confirmacao_personalizada("Confirmar Parada", "Deseja realmente parar a tradução?")
        if resposta:
            # Sinalizar para parar a thread
            self.traducao_ativa = False
            
            # Aguardar a thread terminar (com timeout para não travar)
            if self.thread_traducao and self.thread_traducao.is_alive():
                self.thread_traducao.join(timeout=2.0)
            
            # Atualizar interface
            self.btn_iniciar.configure(state="normal")
            self.btn_parar.configure(state="disabled")
            self.label_status_bar.configure(text="⏹️ Tradução parada")
            self.log_atividade("Tradução parada pelo usuário")
            
            # Resetar progresso
            self.progress_bar.set(0)
            self.label_progress.configure(text="0%")
    
    def exportar_resultado(self, formato):
        """Exporta o resultado da tradução"""
        if not hasattr(self, 'df_full_path') or not self.df_full_path:
            messagebox.showwarning("Aviso", "Nenhum dataset carregado para exportar.")
            return
            
        # Solicitar local de salvamento
        filename = filedialog.asksaveasfilename(
            title="Salvar Resultado Traduzido",
            defaultextension=f".{formato.lower()}",
            filetypes=[
                ("CSV", "*.csv"),
                ("Excel", "*.xlsx"),
                ("SQLite", "*.db")
            ]
        )
        
        if not filename:
            return
            
        try:
            # TODO: Implementar exportação do resultado traduzido
            # Por enquanto, apenas mostrar mensagem
            messagebox.showinfo(
                "Exportação", 
                f"Funcionalidade de exportação será implementada em breve.\n"
                f"Formato selecionado: {formato}\n"
                f"Local: {filename}"
            )
            self.log_atividade(f"Exportação solicitada: {formato} -> {filename}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar: {str(e)}")
            self.log_atividade(f"ERRO na exportação: {str(e)}")
    
    def salvar_configuracoes(self):
        """Salva as configurações atuais"""
        try:
            # Atualizar configurações dos sliders
            self.config['tamanho_lote'] = int(self.slider_lote.get())
            # Converter delay de milissegundos para segundos
            self.config['delay_traducao'] = float(self.slider_delay.get()) / 1000.0
            
            # Salvar em arquivo
            config_file = Path("settings.json")
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
            
            messagebox.showinfo("Sucesso", "Configurações salvas com sucesso!")
            self.log_atividade("Configurações salvas")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar configurações: {str(e)}")
            self.log_atividade(f"ERRO ao salvar configurações: {str(e)}")
    
    def limpar_log(self):
        """Limpa o log de atividades"""
        self.textbox_log.delete("0.0", "end")
        self.log_atividade("Log limpo")
    
    def log_atividade(self, mensagem):
        """Adiciona mensagem ao log de atividades"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {mensagem}\n"
        
        self.textbox_log.insert("end", log_entry)
        self.textbox_log.see("end")
    
    def monitorar_progresso(self):
        """Monitora o progresso da tradução através da fila de mensagens"""
        try:
            # Verificar se há mensagens na fila
            while not self.progress_queue.empty():
                tipo, mensagem = self.progress_queue.get_nowait()
                
                if tipo == "progresso":
                    # Atualizar barra de progresso
                    self.progress_bar.set(mensagem / 100)
                    self.label_progress.configure(text=f"{mensagem:.1f}%")
                    
                elif tipo == "erro":
                    # Mostrar erro
                    self.mostrar_dialogo_personalizado("Erro na Tradução", mensagem, "error")
                    self.log_atividade(f"ERRO: {mensagem}")
                    self.traducao_ativa = False
                    self.btn_iniciar.configure(state="normal")
                    self.btn_parar.configure(state="disabled")
                    self.label_status_bar.configure(text="❌ Erro na tradução")
                    
                elif tipo == "concluido":
                    # Tradução concluída
                    self.mostrar_dialogo_personalizado("Sucesso", "Tradução concluída com sucesso!", "info")
                    self.log_atividade("Tradução concluída")
                    self.traducao_ativa = False
                    self.btn_iniciar.configure(state="normal")
                    self.btn_parar.configure(state="disabled")
                    self.label_status_bar.configure(text="✅ Tradução concluída")
                    self.progress_bar.set(1.0)
                    self.label_progress.configure(text="100%")
            
            # Agendar próxima verificação (mais frequente para melhor responsividade)
            self.root.after(50, self.monitorar_progresso)
            
        except Exception as e:
            self.log_atividade(f"ERRO no monitoramento: {str(e)}")
            # Continuar monitorando mesmo com erro
            self.root.after(50, self.monitorar_progresso)
    
    def run(self):
        """Executa a aplicação"""
        self.root.mainloop()

    def mostrar_dialogo_personalizado(self, titulo, mensagem, tipo="info"):
        """Mostra um diálogo personalizado que combina com o tema da aplicação"""
        # Criar janela de diálogo
        dialog = ctk.CTkToplevel(self.root)
        dialog.title(titulo)
        dialog.geometry("450x300")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Centralizar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"450x300+{x}+{y}")
        
        # Configurar cores baseado no tipo
        if tipo == "info":
            cor_primaria = self.cores['accent']
            cor_secundaria = self.cores['accent_hover']
            icone = "✅"
        elif tipo == "warning":
            cor_primaria = self.cores['warning']
            cor_secundaria = self.cores['warning_hover']
            icone = "⚠️"
        elif tipo == "error":
            cor_primaria = self.cores['danger']
            cor_secundaria = self.cores['danger_hover']
            icone = "❌"
        else:
            cor_primaria = self.cores['accent']
            cor_secundaria = self.cores['accent_hover']
            icone = "ℹ️"
        
        # Frame principal
        main_frame = ctk.CTkFrame(dialog, fg_color=self.cores['secondary'], corner_radius=15)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header com ícone e título
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(20, 15))
        
        # Ícone
        icon_label = ctk.CTkLabel(
            header_frame,
            text=icone,
            font=ctk.CTkFont(size=32),
            text_color=cor_primaria
        )
        icon_label.pack()
        
        # Título
        title_label = ctk.CTkLabel(
            header_frame,
            text=titulo,
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=self.cores['text_primary']
        )
        title_label.pack(pady=(10, 0))
        
        # Mensagem
        message_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        message_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        message_label = ctk.CTkLabel(
            message_frame,
            text=mensagem,
            font=ctk.CTkFont(size=13),
            text_color=self.cores['text_secondary'],
            wraplength=380,
            justify="center"
        )
        message_label.pack()
        
        # Botão OK
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 20))
        
        btn_ok = ctk.CTkButton(
            btn_frame,
            text="OK",
            font=ctk.CTkFont(size=14, weight="bold"),
            height=36,
            corner_radius=8,
            fg_color=cor_primaria,
            hover_color=cor_secundaria,
            text_color=self.cores['text_primary'],
            command=dialog.destroy
        )
        btn_ok.pack()
        
        # Focar no botão OK
        btn_ok.focus_set()
        
        # Bind Enter para fechar
        dialog.bind("<Return>", lambda e: dialog.destroy())
        dialog.bind("<Escape>", lambda e: dialog.destroy())
        
        # Centralizar na tela
        dialog.lift()
        dialog.focus_force()
        
        return dialog

    def mostrar_confirmacao_personalizada(self, titulo, mensagem):
        """Mostra um diálogo de confirmação personalizado (sim/não)"""
        # Variável para armazenar a resposta
        resposta = {"resultado": False}
        
        # Criar janela de diálogo
        dialog = ctk.CTkToplevel(self.root)
        dialog.title(titulo)
        dialog.geometry("500x250")
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)
        
        # Centralizar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (250 // 2)
        dialog.geometry(f"500x250+{x}+{y}")
        
        # Frame principal
        main_frame = ctk.CTkFrame(dialog, fg_color=self.cores['secondary'], corner_radius=15)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header com ícone e título
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(15, 10))
        
        # Ícone de confirmação
        icon_label = ctk.CTkLabel(
            header_frame,
            text="❓",
            font=ctk.CTkFont(size=28),
            text_color=self.cores['accent']
        )
        icon_label.pack()
        
        # Título
        title_label = ctk.CTkLabel(
            header_frame,
            text=titulo,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=self.cores['text_primary']
        )
        title_label.pack(pady=(8, 0))
        
        # Mensagem
        message_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        message_frame.pack(fill="x", padx=20, pady=(0, 15))
        
        message_label = ctk.CTkLabel(
            message_frame,
            text=mensagem,
            font=ctk.CTkFont(size=12),
            text_color=self.cores['text_secondary'],
            wraplength=420,
            justify="center"
        )
        message_label.pack()
        
        # Botões (Sim/Não)
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 15))
        
        # Botão Sim
        btn_sim = ctk.CTkButton(
            btn_frame,
            text="Sim",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=32,
            corner_radius=6,
            fg_color=self.cores['accent'],
            hover_color=self.cores['accent_hover'],
            text_color=self.cores['text_primary'],
            command=lambda: self._confirmar_dialogo(dialog, resposta, True)
        )
        btn_sim.pack(side="left", padx=(0, 8), fill="x", expand=True)
        
        # Botão Não
        btn_nao = ctk.CTkButton(
            btn_frame,
            text="Não",
            font=ctk.CTkFont(size=13, weight="bold"),
            height=32,
            corner_radius=6,
            fg_color=self.cores['glass'],
            hover_color=self.cores['secondary_hover'],
            text_color=self.cores['text_primary'],
            command=lambda: self._confirmar_dialogo(dialog, resposta, False)
        )
        btn_nao.pack(side="left", padx=(8, 0), fill="x", expand=True)
        
        # Focar no botão Sim
        btn_sim.focus_set()
        
        # Bind Enter para Sim, Escape para Não
        dialog.bind("<Return>", lambda e: self._confirmar_dialogo(dialog, resposta, True))
        dialog.bind("<Escape>", lambda e: self._confirmar_dialogo(dialog, resposta, False))
        
        # Centralizar na tela
        dialog.lift()
        dialog.focus_force()
        
        # Aguardar resposta
        dialog.wait_window()
        
        return resposta["resultado"]
    
    def _confirmar_dialogo(self, dialog, resposta, valor):
        """Método auxiliar para confirmar diálogo"""
        resposta["resultado"] = valor
        dialog.destroy()

def main():
    """Função principal da aplicação"""
    app = TradutorCustomTkinterUX()
    app.run()

if __name__ == "__main__":
    main()
